#!/usr/bin/env python3
"""
Convert TradingView Liquidity Data Exporter xlsx to CSV format.

Parses the "List of trades" sheet from the TradingView strategy export,
extracts the pipe-delimited Signal column, and writes a CSV matching the
existing NQ liquidity levels format.

Usage:
    python3 scripts/convert-lt-excel.py <input.xlsx> <output.csv>

Example:
    python3 scripts/convert-lt-excel.py \
        /mnt/c/temp/liquidity/Liquidity_Data_Exporter_CME_MINI_ES1!_2026-02-09_e6313.xlsx \
        backtest-engine/data/liquidity/es/ES_liquidity_levels_1D.csv
"""

import sys
import os
from datetime import datetime, timezone

import openpyxl


def parse_signal(signal_str):
    """Parse pipe-delimited signal: LIQUIDITY|unix_ms|sentiment|l1|l2|l3|l4|l5|id"""
    parts = signal_str.split("|")
    if len(parts) < 8 or parts[0] != "LIQUIDITY":
        return None

    unix_ms = int(parts[1])
    sentiment = parts[2]
    levels = [float(parts[i]) for i in range(3, 8)]

    dt = datetime.fromtimestamp(unix_ms / 1000, tz=timezone.utc)
    dt_str = dt.strftime("%Y-%m-%d %H:%M:%S")

    return {
        "datetime": dt_str,
        "unix_timestamp": unix_ms,
        "sentiment": sentiment,
        "levels": levels,
    }


def convert(input_path, output_path):
    print(f"Loading {input_path}...")
    wb = openpyxl.load_workbook(input_path, read_only=True)
    ws = wb["List of trades"]

    # Find Signal column index
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    signal_idx = headers.index("Signal")

    # Parse all signals, deduplicate by unix timestamp
    seen = {}
    row_count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        signal_str = row[signal_idx]
        if not signal_str or not isinstance(signal_str, str):
            continue
        row_count += 1
        parsed = parse_signal(signal_str)
        if parsed is None:
            continue
        # Deduplicate: keep first occurrence per timestamp
        ts = parsed["unix_timestamp"]
        if ts not in seen:
            seen[ts] = parsed

    wb.close()

    # Sort by timestamp
    records = sorted(seen.values(), key=lambda r: r["unix_timestamp"])

    # Write CSV
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    with open(output_path, "w") as f:
        f.write("datetime,unix_timestamp,sentiment,level_1,level_2,level_3,level_4,level_5\n")
        for r in records:
            levels_str = ",".join(str(v) for v in r["levels"])
            f.write(f"{r['datetime']},{r['unix_timestamp']},{r['sentiment']},{levels_str}\n")

    print(f"Processed {row_count} rows -> {len(records)} unique timestamps")
    print(f"Date range: {records[0]['datetime']} to {records[-1]['datetime']}")
    print(f"Written to {output_path}")


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print(f"Usage: {sys.argv[0]} <input.xlsx> <output.csv>")
        sys.exit(1)
    convert(sys.argv[1], sys.argv[2])
